import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from tkinter import Tk, filedialog, messagebox  # Import messagebox module
from openpyxl.styles import Font

# Função para manter apenas as colunas desejadas
def filter_columns(df):
    desired_columns = ["Full Name", "Project Names", "Leader", "Activities developed in the current week"]
    return df[desired_columns]

# Busca o root onde está o script
root_dir = os.path.dirname(os.path.abspath(__file__))

# Obtém o nome do arquivo Excel a ser convertido
root = Tk()
root.withdraw()
excel_path = filedialog.askopenfilename(initialdir=root_dir, title="Selecione o arquivo .xlsx a ser convertido", filetypes=[("Excel Files", "*.xlsx")])

if not excel_path:
    print('Nenhum arquivo .xlsx selecionado.')
    exit()

# Lê o arquivo Excel no pandas DataFrame
df = pd.read_excel(excel_path, dtype=str)

# Verifica se todas as colunas desejadas estão presentes
desired_columns = ["Full Name", "Project Names", "Leader", "Activities developed in the current week"]
missing_columns = set(desired_columns) - set(df.columns)
if missing_columns:
    # Exibe uma mensagem de erro e encerra o script se faltar alguma coluna
    error_message = f"Missing Values: {', '.join(missing_columns)}"
    messagebox.showerror("Error", error_message)
    exit()

# Filtra as colunas desejadas
df_filtered = filter_columns(df)

# Obtém o nome do arquivo selecionado e usa esse nome para nomear o novo .xlsx
filename = os.path.splitext(os.path.basename(excel_path))[0]
excel_filename = f'{filename}_filtered_{pd.Timestamp.now().strftime("%Y%m%d%H%M")}.xlsx'
excel_filtered_path = os.path.join(root_dir, excel_filename)

# Salva o DataFrame filtrado no novo arquivo Excel
df_filtered.to_excel(excel_filtered_path, index=False)

# Carrega o workbook
workbook = load_workbook(excel_filtered_path)
ws = workbook.active

# Cria uma tabela formatada
table_range = ws.dimensions
table = Table(displayName='Table1', ref=table_range)
style = TableStyleInfo(name="TableStyleMedium1", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
table.tableStyleInfo = style
ws.add_table(table)

# Configuração para definir a cor da fonte do cabeçalho como branca
header_font = Font(color="FFFFFF")
for header_cell in ws[1]:
    header_cell.font = header_font

# Ajusta automaticamente a largura das colunas
for col in ws.columns:
    max_length = 0
    header_length = len(str(col[0].value))  # Get length of header
    for cell in col:
        try:
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length
        except:
            pass
    adjusted_width = max(header_length, max_length) + 2
    column = col[0].column_letter  # Get column name
    ws.column_dimensions[column].width = adjusted_width

# Salva as alterações no arquivo Excel
workbook.save(excel_filtered_path)

print(f'Arquivo Excel filtrado salvo em: {excel_filtered_path}')
